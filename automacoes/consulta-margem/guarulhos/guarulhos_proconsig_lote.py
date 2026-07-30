#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from guarulhos_proconsig_piloto import ProconsigClient


BASE_HEADERS = [
    "linha_origem",
    "nome_planilha",
    "matricula_consultada",
    "status",
    "positivo",
    "nome_confere",
    "mensagem",
    "nome_portal",
    "cpf_portal",
    "regime",
    "cargo_portal",
    "orgao_portal",
    "lotacao_portal",
    "ultima_folha",
    "margem_cartao_beneficios_saque_bruta",
    "margem_cartao_beneficios_saque_liquida",
    "margem_emprestimo_planos_saude_bruta",
    "margem_emprestimo_planos_saude_liquida",
    "margem_cartao_credito_consignado_bruta",
    "margem_cartao_credito_consignado_liquida",
    "linhas_retornadas",
    "href_detalhe",
    "consultado_em",
]


SOURCE_FIELDS = [
    "LOCAL",
    "ADMISSAO",
    "CARGO",
    "COMISSIONAMENTO",
    "REMUNERACAO",
    "DESCONTOS",
    "LIQUIDO",
    "REFERENCIA",
    "CF_DEMITIDO",
    "DEMITIDO",
]


def normalize_text(value):
    text = unicodedata.normalize("NFD", str(value or "").upper())
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_money(value):
    text = str(value or "")
    text = re.sub(r"[^\d,.-]+", "", text)
    if not text:
        return 0.0
    try:
        return float(text.replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def is_positive(result):
    if result.get("status") != "OK":
        return "não"
    keys = [
        "margem_cartao_beneficios_saque_liquida",
        "margem_emprestimo_planos_saude_liquida",
        "margem_cartao_credito_consignado_liquida",
    ]
    return "sim" if any(parse_money(result.get(key)) > 0 for key in keys) else "não"


def load_source(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(header).strip() if header is not None else "" for header in next(iterator)]
    header_index = {header: idx for idx, header in enumerate(headers) if header}
    for required in ["NOME", "COD_FUNC"]:
        if required not in header_index:
            raise SystemExit(f"Coluna obrigatória ausente: {required}")

    rows = []
    for row_number, row_values in enumerate(iterator, 2):
        matricula = row_values[header_index["COD_FUNC"]] if header_index["COD_FUNC"] < len(row_values) else None
        nome = row_values[header_index["NOME"]] if header_index["NOME"] < len(row_values) else None
        if matricula in (None, ""):
            continue
        item = {
            "linha_origem": row_number,
            "nome_planilha": str(nome or "").strip(),
            "matricula_consultada": str(matricula).strip().split(".")[0],
        }
        for field in SOURCE_FIELDS:
            if field in header_index:
                value = row_values[header_index[field]] if header_index[field] < len(row_values) else None
                item[f"planilha_{field.lower()}"] = value.isoformat() if hasattr(value, "isoformat") else value
        rows.append(item)
    return rows


def normalize_result(source, result):
    nome_portal = result.get("nome") or result.get("nome_portal") or ""
    status = result.get("status", "")
    nome_confere = ""
    if status == "OK":
        nome_confere = "sim" if normalize_text(source.get("nome_planilha")) == normalize_text(nome_portal) else "não"
    out = {
        **source,
        "status": status,
        "positivo": "",
        "nome_confere": nome_confere,
        "mensagem": result.get("mensagem", ""),
        "nome_portal": nome_portal,
        "cpf_portal": result.get("cpf", ""),
        "regime": result.get("regime", ""),
        "cargo_portal": result.get("cargo", ""),
        "orgao_portal": result.get("orgao", ""),
        "lotacao_portal": result.get("lotacao", ""),
        "ultima_folha": result.get("ultima_folha", ""),
        "margem_cartao_beneficios_saque_bruta": result.get("margem_cartao_beneficios_saque_bruta", ""),
        "margem_cartao_beneficios_saque_liquida": result.get("margem_cartao_beneficios_saque_liquida", ""),
        "margem_emprestimo_planos_saude_bruta": result.get("margem_emprestimo_planos_saude_bruta", ""),
        "margem_emprestimo_planos_saude_liquida": result.get("margem_emprestimo_planos_saude_liquida", ""),
        "margem_cartao_credito_consignado_bruta": result.get("margem_cartao_credito_consignado_bruta", ""),
        "margem_cartao_credito_consignado_liquida": result.get("margem_cartao_credito_consignado_liquida", ""),
        "linhas_retornadas": result.get("linhas_retornadas", ""),
        "href_detalhe": result.get("href_detalhe", ""),
        "consultado_em": result.get("consultado_em", ""),
    }
    out["positivo"] = is_positive(result) if nome_confere != "não" else "não"
    return out


def load_checkpoint(path):
    if not path.exists():
        return []
    return json.loads(path.read_text("utf-8"))


def save_checkpoint(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), "utf-8")


def write_sheet(workbook, title, rows, headers):
    ws = workbook.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, 1):
        width = max(len(str(header)), *(len(str(ws.cell(row=r, column=idx).value or "")) for r in range(2, ws.max_row + 1)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 55)


def write_xlsx(path, rows):
    extra_headers = []
    for row in rows:
        for key in row:
            if key not in BASE_HEADERS and key not in extra_headers:
                extra_headers.append(key)
    headers = BASE_HEADERS + sorted(extra_headers)
    positive_rows = [row for row in rows if row.get("positivo") == "sim"]
    review_rows = [row for row in rows if row.get("status") != "OK" or row.get("nome_confere") == "não"]
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Todos", rows, headers)
    write_sheet(wb, "Positivos", positive_rows, headers)
    write_sheet(wb, "Revisar", review_rows, headers)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--checkpoint", required=True)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--flush-every", type=int, default=50)
    parser.add_argument("--xlsx-every", type=int, default=1000)
    parser.add_argument("--verbose-every", type=int, default=100)
    parser.add_argument("--request-timeout", type=int, default=15)
    parser.add_argument("--retries", type=int, default=2)
    parser.add_argument("--delay", type=float, default=0.15)
    args = parser.parse_args()

    import os
    cpf = os.environ.get("PROCONSIG_CPF", "")
    password = os.environ.get("PROCONSIG_PASSWORD", "")
    if not cpf or not password:
        raise SystemExit("Configure PROCONSIG_CPF e PROCONSIG_PASSWORD no ambiente.")

    source_rows = load_source(Path(args.input))
    if args.offset:
        source_rows = source_rows[args.offset:]
    if args.limit:
        source_rows = source_rows[:args.limit]

    checkpoint_path = Path(args.checkpoint)
    done_rows = load_checkpoint(checkpoint_path)
    done_keys = {str(row.get("matricula_consultada")) for row in done_rows}

    client = ProconsigClient(cpf, password, timeout=args.request_timeout)
    client.login()

    import time
    from datetime import datetime
    for idx, source in enumerate(source_rows, 1):
        matricula = str(source["matricula_consultada"])
        if matricula in done_keys:
            continue
        result = None
        last_error = None
        for attempt in range(1, args.retries + 2):
            try:
                result = client.query_matricula(matricula)
                break
            except Exception as exc:
                last_error = exc
                try:
                    client.login()
                except Exception:
                    pass
                if attempt <= args.retries:
                    time.sleep(max(args.delay, 0.5))
        if result is None:
            result = {"status": "ERRO", "mensagem": str(last_error)}
        result["consultado_em"] = datetime.now().isoformat(timespec="seconds")
        row = normalize_result(source, result)
        done_rows.append(row)
        done_keys.add(matricula)
        if args.verbose_every <= 1 or len(done_rows) % args.verbose_every == 0 or row.get("status") not in {"OK", "NAO_ENCONTRADO"}:
            print(f"{len(done_rows)}/{len(source_rows)} {matricula} {row['status']} positivo={row['positivo']} nome={row['nome_confere']}", flush=True)
        if len(done_rows) % args.flush_every == 0:
            save_checkpoint(checkpoint_path, done_rows)
        if len(done_rows) % args.xlsx_every == 0:
            write_xlsx(Path(args.output), done_rows)
        time.sleep(args.delay)

    save_checkpoint(checkpoint_path, done_rows)
    write_xlsx(Path(args.output), done_rows)


if __name__ == "__main__":
    main()
