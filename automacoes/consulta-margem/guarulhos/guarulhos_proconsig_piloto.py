#!/usr/bin/env python3
import argparse
import html
import os
import re
import socket
import ssl
import time
import urllib.parse
import urllib.request
import http.cookiejar
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://proconsig.com.br"

TARGET_MARGINS = {
    "Margem Cartão de Benefícios (Saque)": (
        "margem_cartao_beneficios_saque_bruta",
        "margem_cartao_beneficios_saque_liquida",
    ),
    "Margem Empréstimo e Planos de Saúde": (
        "margem_emprestimo_planos_saude_bruta",
        "margem_emprestimo_planos_saude_liquida",
    ),
    "Margem Cartão de Crédito Consignado": (
        "margem_cartao_credito_consignado_bruta",
        "margem_cartao_credito_consignado_liquida",
    ),
}

HEADERS = [
    "matricula_consultada",
    "status",
    "mensagem",
    "nome",
    "cpf",
    "regime",
    "cargo",
    "orgao",
    "lotacao",
    "ultima_folha",
    "margem_cartao_beneficios_saque_bruta",
    "margem_cartao_beneficios_saque_liquida",
    "margem_emprestimo_planos_saude_bruta",
    "margem_emprestimo_planos_saude_liquida",
    "margem_cartao_credito_consignado_bruta",
    "margem_cartao_credito_consignado_liquida",
    "linhas_retornadas",
    "href_detalhe",
    "consultado_em",
]


def strip_tags(value):
    text = re.sub(r"<[^>]+>", " ", value or "")
    return html.unescape(re.sub(r"\s+", " ", text).strip())


def only_digits(value):
    return re.sub(r"\D+", "", value or "")


def clean_money(value):
    value = strip_tags(value)
    match = re.search(r"R\$\s*[-\d.,]+", value)
    return match.group(0).replace("\xa0", " ") if match else ""


def extract_field(page, label):
    pattern = re.compile(
        rf"<strong>\s*{re.escape(label)}\s*</strong>\s*([^<]*)",
        re.IGNORECASE,
    )
    match = pattern.search(page)
    return strip_tags(match.group(1)) if match else ""


def parse_rows(page):
    rows = []
    tbody_match = re.search(r"<tbody>([\s\S]*?)</tbody>", page, re.IGNORECASE)
    tbody = tbody_match.group(1) if tbody_match else ""
    row_pattern = re.compile(
        r"<tr[^>]*>\s*"
        r"<td[^>]*>([\s\S]*?)</td>\s*"
        r"<td[^>]*>([\s\S]*?)</td>\s*"
        r"<td[^>]*>([\s\S]*?)</td>\s*"
        r"<td[^>]*>[\s\S]*?<a[^>]+href=\"([^\"]+)\"[\s\S]*?>\s*Selecionar\s*</a>[\s\S]*?</td>\s*"
        r"</tr>",
        re.IGNORECASE,
    )
    for match in row_pattern.finditer(tbody):
        rows.append(
            {
                "cpf": only_digits(strip_tags(match.group(1))),
                "matricula": strip_tags(match.group(2)),
                "nome": strip_tags(match.group(3)),
                "href": html.unescape(match.group(4)),
            }
        )
    return rows


def parse_margins(page):
    margins = {}
    card_pattern = re.compile(
        r"<div class=\"card-header fw-bold d-flex justify-content-between align-items-center\">\s*"
        r"([\s\S]*?)</div>\s*"
        r"<div class=\"card-body\">([\s\S]*?)</div>\s*</div>",
        re.IGNORECASE,
    )
    for match in card_pattern.finditer(page):
        title = strip_tags(match.group(1))
        if title not in TARGET_MARGINS:
            continue
        body = match.group(2)
        gross = clean_money((re.search(r"<strong>\s*Bruta:\s*</strong>\s*([^<]+)", body, re.IGNORECASE) or ["", ""])[1])
        available_match = re.search(
            r"<strong>\s*Dispon[ií]vel:\s*</strong>[\s\S]*?<span[^>]*>\s*([^<]+)\s*</span>",
            body,
            re.IGNORECASE,
        ) or re.search(r"<strong>\s*Dispon[ií]vel:\s*</strong>\s*([^<]+)", body, re.IGNORECASE)
        available = clean_money(available_match.group(1) if available_match else "")
        margins[title] = {"bruta": gross, "liquida": available}
    return margins


class ProconsigClient:
    def __init__(self, cpf, password, timeout=30):
        self.cpf = only_digits(cpf)
        self.password = password
        self.timeout = timeout
        socket.setdefaulttimeout(timeout)
        self.context = ssl._create_unverified_context()
        self.cookies = http.cookiejar.CookieJar()
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookies),
            urllib.request.HTTPSHandler(context=self.context),
        )
        self.opener.addheaders = [("User-Agent", "Mozilla/5.0"), ("Connection", "close")]

    def open(self, path, data=None):
        url = path if path.startswith("http") else f"{BASE_URL}{path}"
        headers = {}
        payload = None
        if data is not None:
            payload = urllib.parse.urlencode(data).encode()
            headers["Content-Type"] = "application/x-www-form-urlencoded"
        request = urllib.request.Request(url, data=payload, headers=headers)
        with self.opener.open(request, timeout=self.timeout) as response:
            return response.read().decode("utf-8", "replace"), response.url

    def login(self):
        login_path = f"/login?next={urllib.parse.quote(BASE_URL + '/consulta_margem', safe=':/?=&')}"
        self.open(login_path)
        page, url = self.open(login_path, {"cpf": self.cpf, "senha": self.password})
        if "/dashboard" not in url and "Olá," not in page:
            raise RuntimeError("Login no ProConsig falhou.")

    def query_matricula(self, matricula):
        query = urllib.parse.urlencode({"cpf": "", "matricula": matricula, "nome": ""})
        page, _ = self.open(f"/consulta_margem?{query}")
        plain = strip_tags(page)
        if "servidor ineleg" in plain.lower() or "inelegível" in plain.lower() or "inelegivel" in plain.lower():
            return {"status": "INELEGIVEL", "mensagem": "Servidor inelegível pelo cargo.", "linhas_retornadas": 0}
        rows = parse_rows(page)
        if not rows:
            message = "Nenhum servidor encontrado."
            return {"status": "NAO_ENCONTRADO", "mensagem": message, "linhas_retornadas": 0}
        exact = [row for row in rows if str(row["matricula"]).strip() == str(matricula).strip()]
        if len(exact) != 1:
            return {
                "status": "AMBIGUO" if len(exact) > 1 else "SEM_MATRICULA_EXATA",
                "mensagem": f"Busca retornou {len(rows)} linha(s), mas {len(exact)} com matrícula exata.",
                "linhas_retornadas": len(rows),
            }
        selected = exact[0]
        detail, detail_url = self.open(selected["href"])
        margins = parse_margins(detail)
        record = {
            "status": "OK",
            "mensagem": "Consulta realizada.",
            "nome": extract_field(detail, "Nome:") or selected["nome"],
            "cpf": only_digits(extract_field(detail, "CPF:") or selected["cpf"]),
            "regime": extract_field(detail, "Regime:"),
            "cargo": extract_field(detail, "Cargo:"),
            "orgao": extract_field(detail, "Órgão:"),
            "lotacao": extract_field(detail, "Lotação:"),
            "ultima_folha": extract_field(detail, "Última Folha:"),
            "linhas_retornadas": len(rows),
            "href_detalhe": detail_url,
        }
        for title, (gross_key, liquid_key) in TARGET_MARGINS.items():
            record[gross_key] = margins.get(title, {}).get("bruta", "")
            record[liquid_key] = margins.get(title, {}).get("liquida", "")
        return record


def read_matriculas(path):
    text = Path(path).read_text("utf-8")
    return [line.strip() for line in text.splitlines() if line.strip()]


def write_xlsx(rows, output):
    wb = Workbook()
    ws = wb.active
    ws.title = "Guarulhos"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])
    for idx, header in enumerate(HEADERS, 1):
        width = max(len(header), *(len(str(ws.cell(row=r, column=idx).value or "")) for r in range(2, ws.max_row + 1)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 55)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--delay", type=float, default=0.2)
    args = parser.parse_args()

    cpf = os.environ.get("PROCONSIG_CPF", "")
    password = os.environ.get("PROCONSIG_PASSWORD", "")
    if not cpf or not password:
        raise SystemExit("Configure PROCONSIG_CPF e PROCONSIG_PASSWORD no ambiente.")

    client = ProconsigClient(cpf, password)
    client.login()
    rows = []
    matriculas = read_matriculas(args.input)
    for index, matricula in enumerate(matriculas, 1):
        try:
            result = client.query_matricula(matricula)
        except Exception as exc:
            result = {"status": "ERRO", "mensagem": str(exc)}
            try:
                client.login()
            except Exception:
                pass
        result["matricula_consultada"] = matricula
        result["consultado_em"] = datetime.now().isoformat(timespec="seconds")
        rows.append(result)
        print(f"{index}/{len(matriculas)} {matricula} {result.get('status')} {result.get('nome', '')}")
        time.sleep(args.delay)
    write_xlsx(rows, Path(args.output))


if __name__ == "__main__":
    main()
