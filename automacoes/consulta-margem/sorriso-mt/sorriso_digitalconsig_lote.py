#!/usr/bin/env python3
import argparse
import csv
import html
import http.cookiejar
import json
import re
import ssl
import time
import urllib.parse
import urllib.request
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://sistema.digitalconsig.com.br"
SORRISO_ORGAO_ID = "3"

OUTPUT_HEADERS = [
    "linha_origem",
    "matricula_original",
    "matricula_consultada",
    "status",
    "mensagem",
    "nome_planilha",
    "nome_portal",
    "nome_confere",
    "cpf_portal",
    "categoria",
    "secretaria",
    "data_admissao",
    "data_termino_contrato",
    "margem_emprestimo_bruta",
    "margem_emprestimo_disponivel",
    "margem_cartao_bruta",
    "margem_cartao_disponivel",
    "percentual_emprestimo",
    "percentual_cartao",
    "servicos_retornados",
    "consultado_em",
    "planilha_tipo_calculo",
    "planilha_forma_ingresso",
    "planilha_salario_base",
    "planilha_salario_bruto",
    "planilha_descontos",
    "planilha_salario_liquido",
    "planilha_situacao",
    "planilha_estrutura",
    "planilha_mes",
]


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def strip_html(value):
    return clean_text(re.sub(r"<[^>]+>", " ", html.unescape(value or "")))


def mask_cpf(value):
    digits = re.sub(r"\D+", "", str(value or ""))
    if len(digits) != 11:
        return ""
    return f"{digits[:3]}******{digits[-2:]}"


def normalize_name(value):
    import unicodedata

    text = unicodedata.normalize("NFD", str(value or "").upper())
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def safe_cell(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def padded_matricula(value):
    text = str(value or "").strip()
    if not text:
        return ""
    if "." in text:
        text = text.split(".", 1)[0]
    text = re.sub(r"\D+", "", text)
    return text.zfill(10) if text else ""


def extract_hidden_core(page):
    fields = {}
    for name in [
        "__EVENTTARGET",
        "__EVENTARGUMENT",
        "__LASTFOCUS",
        "__VIEWSTATE",
        "__VIEWSTATEGENERATOR",
        "__EVENTVALIDATION",
    ]:
        match = re.search(r'name="%s"[^>]*value="([^"]*)"' % re.escape(name), page)
        fields[name] = html.unescape(match.group(1)) if match else ""
    return fields


def extract_hidden_all(page):
    fields = {}
    for match in re.finditer(r"<input[^>]+>", page, re.IGNORECASE):
        tag = match.group(0)
        type_match = re.search(r'type="([^"]+)"', tag, re.IGNORECASE)
        name_match = re.search(r'name="([^"]+)"', tag, re.IGNORECASE)
        value_match = re.search(r'value="([^"]*)"', tag, re.IGNORECASE)
        field_type = type_match.group(1).lower() if type_match else ""
        if name_match and field_type == "hidden":
            fields[html.unescape(name_match.group(1))] = html.unescape(value_match.group(1)) if value_match else ""
    return fields


def encode_multipart(fields):
    boundary = "----WebKitFormBoundary" + uuid.uuid4().hex[:16]
    parts = []
    for key, value in fields.items():
        parts.append(
            f'--{boundary}\r\nContent-Disposition: form-data; name="{key}"\r\n\r\n{value}\r\n'.encode()
        )
    parts.append(f"--{boundary}--\r\n".encode())
    return boundary, b"".join(parts)


class DigitalConsigSorrisoClient:
    def __init__(self, login, password, timeout=30):
        self.login_value = login
        self.password = password
        self.timeout = timeout
        self.context = ssl._create_unverified_context()
        self.cookies = http.cookiejar.CookieJar()
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookies),
            urllib.request.HTTPSHandler(context=self.context),
        )
        self.opener.addheaders = [
            ("User-Agent", "Mozilla/5.0"),
            ("Connection", "close"),
        ]
        self.consulta_page = None
        self.consulta_url = None

    def open(self, path, data=None, referer=None, multipart=False):
        url = path if path.startswith("http") else f"{BASE_URL}{path}"
        headers = {}
        payload = None
        if referer:
            headers["Referer"] = referer
        if data is not None:
            if multipart:
                boundary, payload = encode_multipart(data)
                headers["Content-Type"] = f"multipart/form-data; boundary={boundary}"
            else:
                payload = urllib.parse.urlencode(data).encode()
                headers["Content-Type"] = "application/x-www-form-urlencoded"
        request = urllib.request.Request(url, data=payload, headers=headers)
        with self.opener.open(request, timeout=self.timeout) as response:
            return response.read().decode("utf-8", "replace"), response.url

    def login(self):
        page, login_url = self.open("/Login.aspx")
        fields = extract_hidden_core(page)
        fields.update({"txtLogin": self.login_value, "txtSenha": self.password, "Entrar": "Entrar"})
        page, selecao_url = self.open("/Login.aspx", fields, referer=login_url)
        if "LoginSelecao.aspx" not in selecao_url:
            raise RuntimeError("Login DigitalConsig falhou ou nao chegou na selecao de orgao.")

        fields = extract_hidden_core(page)
        fields.update({"ctl00$body$ddlOrgao": SORRISO_ORGAO_ID, "ctl00$body$btnVincular": "Vincular"})
        page, inicial_url = self.open("/LoginSelecao.aspx", fields, referer=selecao_url)
        if "Inicial.aspx" not in inicial_url:
            raise RuntimeError("Nao foi possivel vincular Prefeitura Municipal de Sorriso/MT.")

        self.consulta_page, self.consulta_url = self.open("/Orgaos/ConsultaMargem.aspx", referer=inicial_url)
        if "body_txtMatricula" not in self.consulta_page:
            raise RuntimeError("Tela de consulta de margem nao carregou corretamente.")

    def query_matricula(self, matricula):
        if not self.consulta_page or not self.consulta_url:
            self.login()
        fields = extract_hidden_all(self.consulta_page)
        fields.update(
            {
                "ctl00$body$txtCpf": "",
                "ctl00$body$txtMatricula": matricula,
                "ctl00$body$btn_consultar": "Pesquisar",
            }
        )
        page, url = self.open("/Orgaos/ConsultaMargem.aspx", fields, referer=self.consulta_url, multipart=True)
        if "Acesso Negado" in page:
            self.login()
            raise RuntimeError("Acesso negado na consulta; sessao renovada para retry.")
        self.consulta_page = page
        self.consulta_url = url
        return parse_margin_page(page, matricula)


def parse_margin_page(page, matricula):
    rows = []
    for tr in re.findall(r"<tr[\s\S]*?</tr>", page, re.IGNORECASE):
        cells = re.findall(r"<td[^>]*>([\s\S]*?)</td>", tr, re.IGNORECASE)
        if len(cells) < 11:
            continue
        values = [strip_html(cell) for cell in cells]
        if not values or not re.fullmatch(r"\d{11}", re.sub(r"\D+", "", values[0] or "")):
            continue
        rows.append(
            {
                "cpf": re.sub(r"\D+", "", values[0]),
                "matricula": values[1],
                "nome": values[2],
                "categoria": values[3],
                "secretaria": values[4],
                "data_admissao": values[5],
                "data_termino_contrato": values[6],
                "servico": values[7],
                "margem_bruta": values[8],
                "margem_disponivel": values[9],
                "percentual": values[10],
            }
        )

    exact_rows = [row for row in rows if row["matricula"] == matricula]
    if not exact_rows:
        return {
            "status": "NAO_ENCONTRADO",
            "mensagem": "Nenhum retorno com matricula exata.",
            "servicos_retornados": "",
        }

    record = {
        "status": "OK",
        "mensagem": "Consulta realizada.",
        "nome_portal": exact_rows[0]["nome"],
        "cpf_portal": exact_rows[0]["cpf"],
        "categoria": exact_rows[0]["categoria"],
        "secretaria": exact_rows[0]["secretaria"],
        "data_admissao": exact_rows[0]["data_admissao"],
        "data_termino_contrato": exact_rows[0]["data_termino_contrato"],
        "servicos_retornados": ", ".join(row["servico"] for row in exact_rows),
    }
    for row in exact_rows:
        servico = normalize_name(row["servico"])
        if servico == "EMPRESTIMO":
            record["margem_emprestimo_bruta"] = row["margem_bruta"]
            record["margem_emprestimo_disponivel"] = row["margem_disponivel"]
            record["percentual_emprestimo"] = row["percentual"]
        elif servico == "CARTAO CONSIGNADO":
            record["margem_cartao_bruta"] = row["margem_bruta"]
            record["margem_cartao_disponivel"] = row["margem_disponivel"]
            record["percentual_cartao"] = row["percentual"]
    return record


def load_source(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean_text(header) for header in header_row]
    index = {header: idx for idx, header in enumerate(headers) if header}
    if "Matrícula" not in index:
        raise SystemExit("Coluna obrigatoria ausente: Matrícula")

    by_matricula = {}
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        matricula_original = safe_cell(row[index["Matrícula"]]) if index["Matrícula"] < len(row) else ""
        matricula_consultada = padded_matricula(matricula_original)
        if not matricula_consultada:
            continue
        tipo_calculo = safe_cell(row[index.get("Tipo de cálculo", -1)]) if index.get("Tipo de cálculo", -1) >= 0 else ""
        # A linha Mensal e a referencia operacional. Se houver duplicidade, ela vence.
        current = by_matricula.get(matricula_consultada)
        if current and current.get("planilha_tipo_calculo") == "Mensal":
            continue
        item = {
            "linha_origem": row_number,
            "matricula_original": matricula_original,
            "matricula_consultada": matricula_consultada,
            "nome_planilha": safe_cell(row[index.get("Nome completo", -1)]) if index.get("Nome completo", -1) >= 0 else "",
            "planilha_tipo_calculo": tipo_calculo,
            "planilha_forma_ingresso": safe_cell(row[index.get("Forma de ingresso", -1)])
            if index.get("Forma de ingresso", -1) >= 0
            else "",
            "planilha_salario_base": safe_cell(row[index.get("Salário base", -1)]) if index.get("Salário base", -1) >= 0 else "",
            "planilha_salario_bruto": safe_cell(row[index.get("Salário bruto", -1)])
            if index.get("Salário bruto", -1) >= 0
            else "",
            "planilha_descontos": safe_cell(row[index.get("Descontos", -1)]) if index.get("Descontos", -1) >= 0 else "",
            "planilha_salario_liquido": safe_cell(row[index.get("Salário líquido", -1)])
            if index.get("Salário líquido", -1) >= 0
            else "",
            "planilha_situacao": safe_cell(row[index.get("Situação", -1)]) if index.get("Situação", -1) >= 0 else "",
            "planilha_estrutura": safe_cell(row[index.get("Estrutura", -1)]) if index.get("Estrutura", -1) >= 0 else "",
            "planilha_mes": safe_cell(row[index.get("Mês", -1)]) if index.get("Mês", -1) >= 0 else "",
        }
        by_matricula[matricula_consultada] = item
    return list(by_matricula.values())


def load_checkpoint(path):
    if not path.exists():
        return []
    return json.loads(path.read_text("utf-8"))


def save_checkpoint(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), "utf-8")


def write_csv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=OUTPUT_HEADERS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sorriso MT"
    ws.append(OUTPUT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        ws.append([row.get(header, "") for header in OUTPUT_HEADERS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(OUTPUT_HEADERS, 1):
        width = max(len(header), *(len(str(ws.cell(row=r, column=idx).value or "")) for r in range(2, ws.max_row + 1)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 45)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def merge_result(source, result):
    row = {**source}
    row.update(result)
    if row.get("status") == "OK":
        row["nome_confere"] = "sim" if normalize_name(row.get("nome_planilha")) == normalize_name(row.get("nome_portal")) else "nao"
    else:
        row["nome_confere"] = ""
    row["consultado_em"] = datetime.now().isoformat(timespec="seconds")
    return row


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--output-csv", required=True)
    parser.add_argument("--checkpoint", required=True)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--flush-every", type=int, default=20)
    parser.add_argument("--delay", type=float, default=0.25)
    parser.add_argument("--retries", type=int, default=2)
    args = parser.parse_args()

    import os

    login = os.environ.get("DIGITALCONSIG_LOGIN")
    password = os.environ.get("DIGITALCONSIG_PASSWORD")
    if not login or not password:
        raise SystemExit("Configure DIGITALCONSIG_LOGIN e DIGITALCONSIG_PASSWORD no ambiente.")

    source_rows = load_source(Path(args.input))
    if args.offset:
        source_rows = source_rows[args.offset :]
    if args.limit:
        source_rows = source_rows[: args.limit]

    checkpoint_path = Path(args.checkpoint)
    done_rows = load_checkpoint(checkpoint_path)
    done_keys = {row.get("matricula_consultada") for row in done_rows}

    client = DigitalConsigSorrisoClient(login, password)
    client.login()

    for index, source in enumerate(source_rows, 1):
        matricula = source["matricula_consultada"]
        if matricula in done_keys:
            continue
        result = None
        last_error = None
        for attempt in range(1, args.retries + 2):
            try:
                result = client.query_matricula(matricula)
                break
            except Exception as exc:
                last_error = exc
                try:
                    client.login()
                except Exception:
                    pass
                if attempt <= args.retries:
                    time.sleep(max(args.delay, 1.0))
        if result is None:
            result = {"status": "ERRO", "mensagem": str(last_error)}
        row = merge_result(source, result)
        done_rows.append(row)
        done_keys.add(matricula)
        print(
            f"{len(done_rows)}/{len(source_rows)} {matricula} {row.get('status')} "
            f"cpf={mask_cpf(row.get('cpf_portal'))} nome_confere={row.get('nome_confere')}",
            flush=True,
        )
        if len(done_rows) % args.flush_every == 0:
            save_checkpoint(checkpoint_path, done_rows)
            write_csv(Path(args.output_csv), done_rows)
            write_xlsx(Path(args.output_xlsx), done_rows)
        time.sleep(args.delay)

    save_checkpoint(checkpoint_path, done_rows)
    write_csv(Path(args.output_csv), done_rows)
    write_xlsx(Path(args.output_xlsx), done_rows)


if __name__ == "__main__":
    main()
